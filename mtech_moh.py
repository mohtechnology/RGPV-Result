import os
import time
import base64
import requests
from io import BytesIO
from PIL import Image
import pytesseract
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Path to Tesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


# ========== Utility Functions ==========

def decode_base64_captcha(base64_str):
    base64_data = base64_str.split(',')[-1]
    image_data = base64.b64decode(base64_data)
    image = Image.open(BytesIO(image_data)).convert("L")
    return pytesseract.image_to_string(
        image, config='--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    ).strip()


def get_text(soup, element_id):
    tag = soup.find(id=element_id)
    return tag.text.strip() if tag else ""


def save_result_to_excel(soup):
    """Extract and save student info + subject grades from HTML to Excel"""

    # Extract student info
    student_info = {
        "name": get_text(soup, "ctl00_ContentPlaceHolder1_lblNameGrading"),
        "roll": get_text(soup, "ctl00_ContentPlaceHolder1_lblRollNoGrading"),
        "program": get_text(soup, "ctl00_ContentPlaceHolder1_lblProgramGrading"),
        "branch": get_text(soup, "ctl00_ContentPlaceHolder1_lblBranchGrading"),
        "semester": get_text(soup, "ctl00_ContentPlaceHolder1_lblSemesterGrading"),
        "status": get_text(soup, "ctl00_ContentPlaceHolder1_lblStatusGrading"),
        "session": get_text(soup, "ctl00_ContentPlaceHolder1_lblSession"),
        "result_desc": get_text(soup, "ctl00_ContentPlaceHolder1_lblResultNewGrading"),
        "sgpa": get_text(soup, "ctl00_ContentPlaceHolder1_lblSGPA"),
        "cgpa": get_text(soup, "ctl00_ContentPlaceHolder1_lblcgpa")
    }

    # Extract subjects
    subjects = {}
    tables = soup.find_all("table", class_="gridtable")

    for table in tables:
        rows = table.find_all("tr")
        # Subject tables have 4 <td>
        if rows and len(rows[0].find_all("td")) == 4:
            cols = [td.text.strip() for td in rows[0].find_all("td")]
            if cols[0] and cols[3]:  # Ensure valid subject entry
                subjects[cols[0]] = cols[3]

    if not student_info["name"]:
        print("⚠ No valid result found in the HTML.")
        return

    # ===== Excel Handling =====
    file_name = "RGPV_Result.xlsx"
    sheet_name = "Results"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Designed By Moh Technology"])
        header = ["S.No", "Name", "Roll No", "Program", "Branch", "Semester", "Status", "Session"]
        header += list(subjects.keys()) + ["Result Description", "SGPA", "CGPA"]
        ws.append(header)

    # Ensure subject columns exist
    existing_headers = [cell.value for cell in ws[2]]
    for subject in subjects:
        if subject not in existing_headers:
            ws.cell(row=2, column=len(existing_headers) + 1).value = subject
            existing_headers.append(subject)

    serial = ws.max_row - 1
    row_data = [serial] + list(student_info.values())[:7]

    # Fill subjects
    for sub in existing_headers[8:-3]:
        row_data.append(subjects.get(sub, ""))

    # Append final result fields
    row_data += [student_info["result_desc"], student_info["sgpa"], student_info["cgpa"]]

    while len(row_data) < len(existing_headers):
        row_data.append("")

    ws.append(row_data)

    # Auto-adjust column width
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(file_name)
    print(f"✅ Grades for {student_info['name']} ({student_info['roll']}) saved successfully.")


# ========== Fetch and Parse Result ==========

def fetch_result(program, enrollment_no, semester, grading=True):
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 20)

    try:
        for attempt in range(1, 4):
            driver.get("https://result.rgpv.ac.in/Result/ProgramSelect.aspx")
            wait.until(EC.element_to_be_clickable((By.ID, f"radlstProgram_{program - 1}"))).click()
            wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtrollno")))

            print(f"\n🔄 Attempt #{attempt} for {enrollment_no}")

            driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtrollno").send_keys(enrollment_no)

            # Select semester
            sem_dropdown = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_drpSemester")
            for option in sem_dropdown.find_elements(By.TAG_NAME, "option"):
                if option.get_attribute("value") == semester:
                    option.click()
                    break

            # Select grading/non-grading
            grading_id = "ctl00_ContentPlaceHolder1_rbtnlstSType_0" if grading else "ctl00_ContentPlaceHolder1_rbtnlstSType_1"
            driver.find_element(By.ID, grading_id).click()

            # CAPTCHA
            captcha_img = wait.until(EC.presence_of_element_located((By.XPATH, '//img[contains(@src, "CaptchaImage")]')))
            captcha_url = captcha_img.get_attribute("src")
            response = requests.get(captcha_url)
            captcha_img = Image.open(BytesIO(response.content)).convert("L")
            captcha_text = pytesseract.image_to_string(
                captcha_img, config='--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
            ).strip()

            print(f"🔍 CAPTCHA: {captcha_text}")
            driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TextBox1").send_keys(captcha_text)
            time.sleep(4)
            driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_btnviewresult").click()
            time.sleep(1)

            try:
                alert = driver.switch_to.alert
                print("⚠ Alert:", alert.text)
                alert.accept()
                continue
            except:
                print("✅ Result page loaded successfully.")
                with open("result.html", "w", encoding="utf-8") as f:
                    f.write(driver.page_source)
                break

    except TimeoutException as e:
        print("⛔ Timeout:", e)
    except Exception as e:
        print("❌ Error:", e)
    finally:
        try:
            with open("result.html", "r", encoding="utf-8") as f:
                soup = BeautifulSoup(f, "html.parser")
            save_result_to_excel(soup)
        except Exception as err:
            print("⚠ Unable to save result:", err)
        driver.quit()


def fetch_range(program, prefix, start, end, semester, grading=True):
    for num in range(start, end + 1):
        enr_no = f"{prefix}{str(num).zfill(2)}"
        print(f"\n📘 Fetching result for: {enr_no}")
        fetch_result(program, enr_no, semester, grading)


# =========================
if __name__ == "__main__":
    fetch_range(program=8, prefix="0805cs24me", start=1, end=20, semester="2", grading=True)
