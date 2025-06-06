from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException
from PIL import Image
import pytesseract
import time
import base64
import requests
from io import BytesIO


tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Update if needed
pytesseract.pytesseract.tesseract_cmd = tesseract_cmd


# ===== HELPER FUNCTIONS =====
def decode_captcha_from_base64(img_src_base64):
    # Remove the prefix and decode base64
    base64_data = img_src_base64.split(',')[-1]
    img_data = base64.b64decode(base64_data)
    image = Image.open(BytesIO(img_data)).convert("L")  # Grayscale
    captcha_text = pytesseract.image_to_string(image, config='--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789').strip()
    return captcha_text


def store_result():
    from bs4 import BeautifulSoup
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    import os

    # Read HTML result file (replace with response.text in actual use)
    with open("result.html", "r", encoding="utf-8") as file:
        soup = BeautifulSoup(file, "html.parser")

    def get_text_by_id(soup, id_):
        tag = soup.find(id=id_)
        return tag.text.strip() if tag else ""

    # Extract main details
    name = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblNameGrading")
    roll = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblRollNoGrading")
    program = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblProgramGrading")
    branch = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblBranchGrading")
    semester = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblSemesterGrading")
    status = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblStatusGrading")
    session = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblSession")
    result_desc = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblResultNewGrading")
    sgpa = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblSGPA")
    cgpa = get_text_by_id(soup, "ctl00_ContentPlaceHolder1_lblcgpa")

    # Extract subject grades
    subjects = {}
    if name:  # if result found
        subject_tables = soup.find_all("table", class_="gridtable")
        for table in subject_tables:
            rows = table.find_all("tr")
            for row in rows:
                cols = row.find_all("td")
                if len(cols) == 4:
                    subject = cols[0].text.strip()
                    grade = cols[3].text.strip() 
                    subjects[subject] = grade
    else:
        name = roll = program = branch = semester = status = session = result_desc = sgpa = cgpa = "Not Found"
        subjects = {}

    # Excel file
    file_name = "RGPV_Grades_Clean.xlsx"
    sheet_name = "Results"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Designed By Moh Technology"])  # First row header
        # Header row
        header = ["S.No", "Name", "Roll No", "Program", "Branch", "Semester", "Status", "Session"]
        header += list(subjects.keys())
        header += ["Result Description", "SGPA", "CGPA"]
        ws.append(header)

    # Ensure all subjects are in headers
    existing_headers = [cell.value for cell in ws[2]]  # row 2 is header
    for sub in subjects:
        if sub not in existing_headers:
            ws.cell(row=2, column=len(existing_headers) + 1).value = sub
            existing_headers.append(sub)

    # Serial number based on actual data rows (after header)
    serial_no = ws.max_row - 2
    row = [serial_no, name, roll, program, branch, semester, status, session]

    # Subject grades in same order as headers
    for col_name in existing_headers[8:-3]:
        row.append(subjects.get(col_name, ""))

    # Summary
    row += [result_desc, sgpa, cgpa]

    # Fill blanks if any
    while len(row) < len(existing_headers):
        row.append("")

    ws.append(row)

    # Adjust column widths
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
        max_length = max(len(str(cell.value) if cell.value is not None else "") for cell in col_cells)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save file
    wb.save(file_name)
    print(f"✅ Grades saved successfully in '{file_name}' with width adjusted and title row.")


# ===== MAIN FUNCTION =====
def fetch_result(enrollment_no, semester, grading):
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 20)

    try:
        # Step 1: Go to Program Select
        driver.get("https://result.rgpv.ac.in/Result/ProgramSelect.aspx")

        # Step 2: Select B.Tech.
        wait.until(EC.element_to_be_clickable((By.ID, "radlstProgram_1"))).click()

        # Step 3: Wait for Result Page to Load
        wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtrollno")))

        for attempt in range(1,2):
            print(f"\nAttempt #{attempt}")

            # Fill Enrollment Number
            driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtrollno").clear()
            driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtrollno").send_keys(enrollment_no)

            # Select Semester
            sem_select = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_drpSemester")
            for option in sem_select.find_elements(By.TAG_NAME, "option"):
                if option.get_attribute("value") == semester:
                    option.click()
                    break

            # Select Grading or Non-Grading
            grading_id = "ctl00_ContentPlaceHolder1_rbtnlstSType_0" if grading else "ctl00_ContentPlaceHolder1_rbtnlstSType_1"
            driver.find_element(By.ID, grading_id).click()

            # Wait for CAPTCHA image
            captcha_img = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="ctl00_ContentPlaceHolder1_pnlCaptcha"]//img')))
            captcha_src = captcha_img.get_attribute("src")

            # Fetch and decode CAPTCHA
            response = requests.get(captcha_src)
            img = Image.open(BytesIO(response.content)).convert("L")
            captcha_text = pytesseract.image_to_string(img, config='--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789').strip()
            print("Decoded Captcha:", repr(captcha_text))

            # Fill CAPTCHA
            captcha_input = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_TextBox1")
            captcha_input.clear()
            captcha_input.send_keys(captcha_text)
            time.sleep(3) 

            # Click Submit
            driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_btnviewresult").click()

            time.sleep(0.5)  # Let result load or alert appear

            # Check for alert
            try:
                alert = driver.switch_to.alert
                print("Alert detected:", alert.text)
                alert.accept()
                continue  # Retry
            except:
                pass

            # Result appears — Save or print
            print("✅ Result page reached successfully.")
            with open("result.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("✔ Result saved to result.html")
            store_result()

    except TimeoutException as e:
        print("⛔ Timeout error:", e)
    except Exception as e:
        print("❌ Error:", e)
    finally:
        driver.quit()

def fetch_results_for_range(prefix, start, end, semester, grading):
    for i in range(start, end + 1):
        enrollment_number = f"{prefix}{str(i).zfill(3)}"
        print(f"Fetching for {enrollment_number}")
        fetch_result(enrollment_number, semester, grading)


prefix = "0805CS"
start = 241001
end = 241010
semester = 1
grading = True

fetch_results_for_range(prefix, start, end, semester, grading)